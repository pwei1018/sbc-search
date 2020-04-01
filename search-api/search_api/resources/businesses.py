# Copyright © 2020 Province of British Columbia
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

import datetime
from http import HTTPStatus
from tempfile import NamedTemporaryFile

from flask import Blueprint, request, jsonify, send_from_directory
from openpyxl import Workbook

from search_api.auth import jwt, authorized
from search_api.models import (
    Corporation,
    CorpName,
    Office,
    _get_corporation_search_results,
    _merge_corpparty_search_addr_fields,
    _normalize_addr,
    _format_office_typ_cd
)
from search_api.utils.utils import convert_to_snake_case


API = Blueprint('BUSINESSES_API', __name__, url_prefix='/api/v1/businesses')


@API.route('/')
@jwt.requires_auth
def corporation_search():
    account_id = request.headers.get("X-Account-Id")
    if not authorized(jwt, account_id):
        return jsonify({'message': 'User is not authorized to access Director Search'}), HTTPStatus.UNAUTHORIZED

    args = request.args
    results = _get_corporation_search_results(args)

    # Pagination
    page = int(args.get("page")) if "page" in args else 1
    results = results.paginate(int(page), 50, False)

    corporations = []
    for row in results.items:
        result_dict = {}

        result_fields = ['corpNum', 'corpNme', 'recognitionDts', 'corpTypCd', 'stateTypCd', 'postalCd']

        result_dict = {key: getattr(row, convert_to_snake_case(key)) for key in result_fields}
        result_dict['addr'] = _merge_corpparty_search_addr_fields(row)

        corporations.append(result_dict)

    return jsonify({'results': corporations})


@API.route('/export/')
@jwt.requires_auth
def corporation_search_export():
    account_id = request.headers.get("X-Account-Id")
    if not authorized(jwt, account_id):
        return jsonify({'message': 'User is not authorized to access Director Search'}), HTTPStatus.UNAUTHORIZED

    # Query string arguments
    args = request.args

    # Fetching results
    results = _get_corporation_search_results(args)

    # Exporting to Excel
    wb = Workbook()

    export_dir = "/tmp"
    with NamedTemporaryFile(mode='w+b', dir=export_dir, delete=True):

        sheet = wb.active

        # Sheet headers (first row)
        _ = sheet.cell(column=1, row=1, value="Inc/Reg #")
        _ = sheet.cell(column=2, row=1, value="Entity Type")
        _ = sheet.cell(column=3, row=1, value="Company Name")
        _ = sheet.cell(column=4, row=1, value="Incorporated")
        _ = sheet.cell(column=5, row=1, value="Company Status")
        _ = sheet.cell(column=6, row=1, value="Company Address")
        _ = sheet.cell(column=7, row=1, value="Postal Code")

        for index, row in enumerate(results, 2):
            # Corporation.corp_num
            _ = sheet.cell(column=1, row=index, value=row.corp_num)
            # Corporation.corp_typ_cd
            _ = sheet.cell(column=2, row=index, value=row.corp_typ_cd)
            # CorpName.corp_nme
            _ = sheet.cell(column=3, row=index, value=row.corp_nme)
            # Corporation.recognition_dts
            _ = sheet.cell(column=4, row=index, value=row.recognition_dts)
            # CorpOpState.state_typ_cd
            _ = sheet.cell(column=5, row=index, value=row.state_typ_cd)
            # Address.addr_line_1, Address.addr_line_2, Address.addr_line_3
            _ = sheet.cell(column=6, row=index, value=_merge_corpparty_search_addr_fields(row))
            # Address.postal_cd
            _ = sheet.cell(column=7, row=index, value=row.postal_cd)

        current_date = datetime.datetime.strftime(datetime.datetime.now(), "%Y-%m-%d %H:%M:%S")
        filename = "Corporation Search Results {date}.xlsx".format(date=current_date)
        full_filename_path = "{dir}/{filename}".format(dir=export_dir, filename=filename)
        wb.save(filename=full_filename_path)

        return send_from_directory(export_dir, filename, as_attachment=True)


@API.route('/<id>')
@jwt.requires_auth
def corporation(id):
    account_id = request.headers.get("X-Account-Id")
    if not authorized(jwt, account_id):
        return jsonify({'message': 'User is not authorized to access Director Search'}), HTTPStatus.UNAUTHORIZED

    corp = Corporation.get_corporation_by_id(id)
    offices = Office.get_offices_by_corp_id(id)
    names = CorpName.get_corp_name_by_corp_id(id)

    output = {}
    output['corpNum'] = corp.corp_num
    output['transitionDt'] = corp.transition_dt
    output['offices'] = []
    for office in offices:
        output['offices'].append({
            'deliveryAddr': _normalize_addr(office.delivery_addr_id),
            'mailingAddr': _normalize_addr(office.mailing_addr_id),
            'officeTypCd': _format_office_typ_cd(office.office_typ_cd),
            'emailAddress': office.email_address
        })

    output['adminEmail'] = corp.admin_email

    output['NAMES'] = []
    for row in names:
        output['NAMES'].append({
            'name': row.corp_nme
        })

    return jsonify(output)
